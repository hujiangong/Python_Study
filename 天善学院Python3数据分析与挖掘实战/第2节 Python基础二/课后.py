# 读写2003 Excel
import xlrd
import xlwt
# 读写2007 Excel
import openpyxl


def write07Excel(path):

    wb = openpyxl.Workbook()
    sheet=wb.active
    sheet.title='2007测试表'

    value = [["名称", "价格", "出版社", "语言"],
             ["如何高效读懂一本书", "22.3", "机械工业出版社", "中文"],
             ["暗时间", "32.4", "人民邮电出版社", "中文"],
             ["拆掉思维里的墙", "26.7", "机械工业出版社", "中文"]]

    for i in range(0,4):
        for j in range(0,len(value[i])):
            sheet.cell(row=i+1,column=j+1, value=str(value[i][j]))

    wb.save(path)
    print('写入成功！')

def read07Excel(path):
    wb=openpyxl.load_workbook(path)
    sheet=wb.get_sheet_by_name('2007测试表')

    for row in sheet.rows:
        for cell in row:
            print(cell.value,"\t",end="")
        print()

def write03Excel(path):
    wb = xlwt.Workbook()
    sheet = wb.add_sheet("2003测试表")
    value = [["名称", "价格", "出版社", "语言"],
             ["如何高效读懂一本书", "22.3", "机械工业出版社", "中文"],
             ["暗时间", "32.4", "人民邮电出版社", "中文"],
             ["拆掉思维里的墙", "26.7", "机械工业出版社", "中文"]]
    for i in range(0, 4):
        for j in range(0, len(value[i])):
            sheet.write(i, j, value[i][j])
    wb.save(path)
    print("写入数据成功！")

def read03Excel(path):
    workbook = xlrd.open_workbook(path)
    sheets = workbook.sheet_names()
    worksheet = workbook.sheet_by_name(sheets[0])
    for i in range(0, worksheet.nrows):
        row = worksheet.row(i)
        for j in range(0, worksheet.ncols):
            print(worksheet.cell_value(i, j), "\t", end="")
        print()

def readLengthOfSheets(path):
    workbook = xlrd.open_workbook(path)
    sheet_count=len(workbook.sheets())
    for sheet in workbook.sheets():
        print(sheet.name)
# write07Excel('2007.xlsx')
read07Excel('2007.xlsx')
# write03Excel('2003.xlsx')
# read03Excel('2003.xlsx')
#readLengthOfSheets('2003.xlsx')

# 作业
#print (sheet1.name, sheet1.nrows, sheet1.ncols)
#workbookAll = xlrd.open_workbook('all.xlsx')
# workbook1 = xlrd.open_workbook('2003.xlsx')
# workbook2 = xlrd.open_workbook('2007.xlsx')
# workbook1_sheet={}
# for sheet in workbook1.sheets():
#     workbook1_sheet[sheet.name]={}
#     print('1:',workbook1_sheet)
#     for row_num in range (0, sheet.nrows):
#         workbook1_sheet[sheet.name][str(row_num)]={}
#         print('2:',workbook1_sheet)
#         for cell_num in range(0,sheet.ncols):
#             #workbook1_sheet[sheet.name][str(cell_num)][str(sheet.ncols)]={}
#             workbook1_sheet[sheet.name][str(row_num)][str(cell_num)]=sheet.cell_value(row_num,cell_num)
#             print('3:',workbook1_sheet)